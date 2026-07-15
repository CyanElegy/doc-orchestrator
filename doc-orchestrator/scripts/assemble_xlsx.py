#!/usr/bin/env python3
"""
assemble_xlsx.py — Assemble a .xlsx from skeleton + content + template.

Usage:
    python3 assemble_xlsx.py skeleton.json content.json <template.xlsx> --output <output.xlsx>

Steps:
    1. Copy template → output (inherit styles, formulas, merged cells)
    2. Load skeleton.json and content.json
    3. Build a content lookup by unit id
    4. For each sheet unit in skeleton:
       - Clear existing data cells (preserve formulas and summary rows)
       - Match content data to template column headers
       - Write data rows with template cell styles
    5. Save

Requires: openpyxl
"""

from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------

def assemble(
    skeleton_path: Path,
    content_path: Path,
    template_path: Path,
    output_path: Path,
) -> None:
    """Main assembly workflow."""

    # -- Validate inputs -------------------------------------------------------
    for p, label in [
        (skeleton_path, "Skeleton JSON"),
        (content_path, "Content JSON"),
        (template_path, "Template"),
    ]:
        if not p.exists():
            print(f"Error: {label} not found: {p}", file=sys.stderr)
            sys.exit(1)

    # -- Load data files -------------------------------------------------------
    try:
        skeleton = json.loads(skeleton_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"Error: Invalid skeleton JSON: {exc}", file=sys.stderr)
        sys.exit(1)

    try:
        content_data = json.loads(content_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"Error: Invalid content JSON: {exc}", file=sys.stderr)
        sys.exit(1)

    # Build content lookup
    content_lookup: dict[str, Any] = {}
    for unit in content_data.get("units", []):
        unit_id = unit.get("id", "")
        unit_content = unit.get("content", "")
        if unit_id:
            # Try to parse as JSON array for table data
            if isinstance(unit_content, str) and unit_content.strip().startswith("["):
                try:
                    content_lookup[unit_id] = json.loads(unit_content)
                except json.JSONDecodeError:
                    content_lookup[unit_id] = unit_content
            else:
                content_lookup[unit_id] = unit_content

    # -- Copy template to output -----------------------------------------------
    shutil.copy2(str(template_path), str(output_path))
    wb = load_workbook(str(output_path))

    # -- Process each sheet ----------------------------------------------------
    for unit in skeleton.get("units", []):
        if unit.get("element") != "sheet":
            continue

        unit_id = unit.get("id", "")
        sheet_name = unit.get("sheet_name", "")
        headers = unit.get("headers", [])
        header_row = unit.get("header_row", 1)

        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in template, skipping", file=sys.stderr)
            continue

        ws = wb[sheet_name]

        # Build header → column index map
        header_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val is not None:
                key = str(val).strip()
                header_map[key] = col_idx

        # Unmerge data area cells to avoid write errors
        merged_to_unmerge: list[str] = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row > header_row:
                merged_to_unmerge.append(str(merged_range))
        for range_str in merged_to_unmerge:
            ws.unmerge_cells(range_str)

        # Clear data area (preserve formulas)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if _is_summary_row(ws, row_idx, ws.max_column):
                continue
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not _is_formula(cell.value):
                    cell.value = None

        # Get content data
        table_data = content_lookup.get(unit_id)
        if table_data is None:
            continue

        # Parse into rows
        if isinstance(table_data, str):
            table_data = [[table_data]]
        if not isinstance(table_data, list):
            continue
        if len(table_data) == 0:
            continue

        if isinstance(table_data[0], list):
            rows_data: list[list[Any]] = table_data
        else:
            # Single row: list of values
            rows_data = [table_data]

        # Determine style source row
        data_start_row = header_row + 1
        safe_row = min(data_start_row, ws.max_row)
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=safe_row, column=col_idx).value is None:
                safe_row = max(1, header_row - 1)
                break

        # Write data
        for i, row_data in enumerate(rows_data):
            current_row = data_start_row + i

            if isinstance(row_data, dict):
                # Dict: keys match header names
                unmatched: list[str] = []
                for key, value in row_data.items():
                    if key in header_map:
                        col_idx = header_map[key]
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = value
                        if safe_row >= 1:
                            _copy_style(
                                ws.cell(row=safe_row, column=col_idx), cell
                            )
                    else:
                        unmatched.append(key)
                if unmatched:
                    print(
                        f"Warning: Sheet '{sheet_name}', row {i + 1}: "
                        f"keys not found in template headers: {unmatched}",
                        file=sys.stderr,
                    )
            elif isinstance(row_data, list):
                # List: values in column order
                for col_idx, value in enumerate(row_data):
                    excel_col = col_idx + 1  # 1-indexed
                    if excel_col <= ws.max_column:
                        cell = ws.cell(row=current_row, column=excel_col)
                        cell.value = value
                        if safe_row >= 1:
                            _copy_style(
                                ws.cell(row=safe_row, column=excel_col), cell
                            )

    # -- Save ------------------------------------------------------------------
    wb.save(str(output_path))
    print(f"Done! → {output_path}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_formula(value) -> bool:
    """Check if a value is an Excel formula."""
    return isinstance(value, str) and value.startswith("=")


def _is_summary_row(ws, row_idx: int, max_col: int) -> bool:
    """Check if a row contains formulas or summary keywords."""
    keywords = {"合计", "总计", "小计", "sum", "total", "subtotal"}
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is None:
            continue
        if _is_formula(val):
            return True
        if isinstance(val, str) and any(kw in val.lower() for kw in keywords):
            return True
    return False


def _copy_style(source_cell, target_cell) -> None:
    """Copy cell style from source to target."""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    if len(args) < 3:
        print(
            "Usage: python3 assemble_xlsx.py skeleton.json content.json "
            "<template.xlsx> --output <output.xlsx>",
            file=sys.stderr,
        )
        sys.exit(1)

    output_path: Optional[Path] = None
    i = 3
    while i < len(args):
        if args[i] == "--output" and i + 1 < len(args):
            output_path = Path(args[i + 1])
            i += 2
        else:
            i += 1

    if output_path is None:
        print("Error: --output is required", file=sys.stderr)
        sys.exit(1)

    assemble(
        Path(args[0]),
        Path(args[1]),
        Path(args[2]),
        output_path,
    )


if __name__ == "__main__":
    main()
