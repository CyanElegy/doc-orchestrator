#!/usr/bin/env python3
"""assemble_xlsx.py — 从结构化数据 + 模板生成 .xlsx 文件。

用法: python3 assemble_xlsx.py <data.yaml> <template.xlsx> <output.xlsx>

data.yaml 格式:
  sheets:
    工作表名:
      - 列A的值: 值
        列B的值: 值
      - ...

步骤:
  1. 复制模板 → 继承样式、列宽、合并单元格、公式
  2. 清空数据单元格（保留公式和汇总行）
  3. 按 data.yaml 键名匹配列标题填充数据
  4. 保留模板单元格样式（字体、对齐、边框等）

依赖: openpyxl。如未安装，打印错误并退出。
"""

import sys
import os
from pathlib import Path

try:
    import yaml
except ImportError:
    print("pyyaml 未安装。请执行: pip install pyyaml")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 未安装。请执行: pip install openpyxl")
    sys.exit(1)


def _find_header_row(ws, max_col: int, max_check_rows: int = 20) -> int:
    """在第一个 max_check_rows 行内定位标题行。

    标题行定义为包含非空单元格的行，假设其值是该列的键。
    如果未找到则返回 1。
    """
    for row_idx in range(1, max_check_rows + 1):
        non_empty = 0
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                non_empty += 1
        if non_empty >= max(1, max_col // 2):
            return row_idx
    return 1


def _build_header_map(ws, header_row: int, max_col: int) -> dict[str, int]:
    """将标题文本映射到列索引。"""
    header_map = {}
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value is not None:
            key = str(cell.value).strip()
            header_map[key] = col_idx
    return header_map


def _copy_cell_style(source_cell, target_cell):
    """从源单元格复制样式到目标单元格。"""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()


def _is_formula(value) -> bool:
    """检查一个值是否包含 Excel 公式语法。"""
    return isinstance(value, str) and value.startswith("=")


def _is_summary_row(ws, row_idx: int, max_col: int) -> bool:
    """如果一行中包含公式或汇总关键词，则判定为汇总行。"""
    keywords = {"合计", "总计", "小计", "sum", "total", "subtotal"}
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is None:
            continue
        if _is_formula(val):
            return True
        if isinstance(val, str) and any(kw in val.lower() for kw in keywords):
            return True
    return False


def _clear_data_area(ws, header_row: int, max_row: int, max_col: int):
    """清空标题行下方的所有非公式非汇总数据。"""
    for row_idx in range(header_row + 1, max_row + 1):
        if _is_summary_row(ws, row_idx, max_col):
            continue
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not _is_formula(cell.value):
                cell.value = None


def assemble_xlsx(data_path: str, template_path: str, output_path: str) -> str:
    """主入口：从 data.yaml + template.xlsx → output.xlsx。"""
    if not os.path.isfile(data_path):
        print(f"错误: 数据文件不存在: {data_path}")
        sys.exit(1)

    if not os.path.isfile(template_path):
        print(f"错误: 模板文件不存在: {template_path}")
        sys.exit(1)

    with open(data_path, "r", encoding="utf-8") as f:
        try:
            data = yaml.safe_load(f)
        except yaml.YAMLError as e:
            print(f"错误: YAML 解析失败 ({data_path}): {e}")
            sys.exit(1)

    if data is None:
        print("警告: data.yaml 为空，生成的文件将仅包含模板内容。")
        data = {}

    wb = load_workbook(template_path)

    sheets_data = data.get("sheets", data)

    for sheet_name, rows in sheets_data.items():
        if sheet_name not in wb.sheetnames:
            print(f"警告: 模板中未找到工作表 '{sheet_name}'，跳过。")
            continue

        ws = wb[sheet_name]
        max_col = ws.max_column
        max_row = ws.max_row

        header_row = _find_header_row(ws, max_col)
        header_map = _build_header_map(ws, header_row, max_col)

        _clear_data_area(ws, header_row, max_row, max_col)

        if not isinstance(rows, list):
            print(f"警告: 工作表 '{sheet_name}' 的数据不是列表格式，跳过。")
            continue

        data_start_row = header_row + 1
        style_source_row = data_start_row
        safe_row = min(style_source_row, max_row)
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=safe_row, column=col_idx).value is None:
                safe_row = max(1, header_row - 1)
                break

        for i, row_data in enumerate(rows):
            current_row = data_start_row + i
            if isinstance(row_data, dict):
                for key, value in row_data.items():
                    if key in header_map:
                        col_idx = header_map[key]
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = value
                        if safe_row >= 1:
                            _copy_cell_style(
                                ws.cell(row=safe_row, column=col_idx), cell
                            )

    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("用法: python3 assemble_xlsx.py <data.yaml> <template.xlsx> <output.xlsx>")
        sys.exit(1)

    data_path, template_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]
    result = assemble_xlsx(data_path, template_path, output_path)
    print(f"已生成: {result}")
